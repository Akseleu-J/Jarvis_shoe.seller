import telebot
import urllib.parse
import logging
import csv
import io
import re
import pandas as pd
from abc import ABC, abstractmethod
from datetime import datetime, timedelta
from telebot import types
from apscheduler.schedulers.background import BackgroundScheduler

from config import TOKEN
import database as db

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("jarvis_errors.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)


def build_wa_url(wa_link, brand, category, material, size, partner_id=None):
    partner_note = f" | Заказ через партнера ID: {partner_id}" if partner_id else ""
    text = f"Я от Jarvis, хочу заказать {brand}, {category}, {material}, размера: {size}{partner_note}"
    encoded = urllib.parse.quote(text)
    if wa_link.startswith("http"):
        return f"{wa_link.split('?')[0]}?text={encoded}"
    number = ''.join(filter(str.isdigit, wa_link))
    return f"https://wa.me/{number}?text={encoded}"


def make_reply_keyboard(items, cols=2):
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    row = []
    for item in items:
        row.append(types.KeyboardButton(str(item)))
        if len(row) == cols:
            markup.row(*row)
            row = []
    if row:
        markup.row(*row)
    return markup


def admin_menu_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row(types.KeyboardButton("➕ Добавить товар"))
    markup.row(types.KeyboardButton("📋 Список товаров"))
    markup.row(types.KeyboardButton("📊 Статистика логов"))
    markup.row(types.KeyboardButton("⭐ Отзывы о сервисе"), types.KeyboardButton("📦 Отзывы о товарах"))
    markup.row(types.KeyboardButton("📣 Рассылка всем"))
    markup.row(types.KeyboardButton("📥 Выгрузить в Excel/CSV"))
    return markup


def vendor_menu_keyboard():
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    markup.row(types.KeyboardButton("➕ Добавить товар"))
    markup.row(types.KeyboardButton("📦 Мои товары"))
    markup.row(types.KeyboardButton("📊 Статистика кликов"))
    return markup


ADMIN_ITEM_STEPS = [
    ('ai_brand',    'ai_category', "Шаг 2/12 — Введите *категорию*:"),
    ('ai_category', 'ai_price',    "Шаг 3/12 — Введите *цену* (число в KZT):"),
    ('ai_price',    'ai_material', "Шаг 4/12 — Введите *материал*:"),
    ('ai_material', 'ai_color',    "Шаг 5/12 — Введите *цвет*:"),
    ('ai_color',    'ai_city',     "Шаг 6/12 — Введите *город*:"),
    ('ai_city',     'ai_wa',       "Шаг 7/12 — Введите *номер WhatsApp* (например: 77001234567):"),
    ('ai_wa',       'ai_desc',     "Шаг 8/12 — Введите *описание товара*:"),
    ('ai_desc',     'ai_seller',   "Шаг 9/12 — Введите *имя продавца*:"),
    ('ai_seller',   'ai_photo',    "Шаг 10/12 — Введите *ссылку на фото* (или напишите «нет»):"),
    ('ai_photo',    'ai_sizes',    "Шаг 11/12 — Введите *доступные размеры*\n(например: 36;37;38 или 36-40)\nИли напишите «нет»:"),
    ('ai_sizes',    'ai_partner',  "Шаг 12/12 — Выберите партнёра (введите ID) или напишите «нет»:"),
]

STEP_TO_FIELD = {
    'ai_brand': 'brand', 'ai_category': 'category', 'ai_price': 'price',
    'ai_material': 'material', 'ai_color': 'color', 'ai_city': 'city',
    'ai_wa': 'wa_link', 'ai_desc': 'description', 'ai_seller': 'seller_name',
    'ai_photo': 'photo_url', 'ai_sizes': 'sizes', 'ai_partner': 'partner_id',
}


class BaseHandler(ABC):
    def __init__(self, bot, user_data, scheduler, admin_ids):
        self.bot = bot
        self.user_data = user_data
        self.scheduler = scheduler
        self.admin_ids = admin_ids

    @abstractmethod
    def register(self): pass

    def get_d(self, cid):
        if cid not in self.user_data:
            self.user_data[cid] = {'step': None, 'pending_review': None, 'admin_step': None, 'new_item': {}}
        return self.user_data[cid]

    def is_admin(self, uid): return uid in self.admin_ids

    def notify_admin(self, text):
        for aid in self.admin_ids:
            try: self.bot.send_message(aid, text, parse_mode="Markdown")
            except Exception as e: log.error(f"notify_admin {aid}: {e}")


class ReviewHandler(BaseHandler):
    def register(self):
        @self.bot.message_handler(func=lambda m: self.user_data.get(m.chat.id, {}).get('pending_review') is not None)
        def capture(m): self._capture(m)

    def _capture(self, m):
        cid = m.chat.id
        d = self.get_d(cid)
        rtype = d.get('pending_review')
        if rtype == 'service':
            db.save_review_service(m.from_user.id, m.text.strip())
            self.bot.send_message(cid, "⭐ Спасибо за отзыв о сервисе!", reply_markup=types.ReplyKeyboardRemove())
        elif rtype == 'product':
            db.save_review_product(m.from_user.id, m.text.strip())
            self.bot.send_message(cid, "📦 Спасибо за отзыв о товаре!", reply_markup=types.ReplyKeyboardRemove())
        d['pending_review'] = None

    def schedule_service(self, cid):
        try:
            self.scheduler.add_job(self._ask_service, 'date', run_date=datetime.now()+timedelta(minutes=5),
                                   args=[cid], id=f"svc_{cid}", replace_existing=True)
        except Exception as e: log.error(f"schedule_service: {e}")

    def schedule_product(self, cid):
        try:
            self.scheduler.add_job(self._ask_product, 'date', run_date=datetime.now()+timedelta(hours=24),
                                   args=[cid], id=f"prd_{cid}", replace_existing=True)
        except Exception as e: log.error(f"schedule_product: {e}")

    def cancel_service(self, cid):
        try: self.scheduler.remove_job(f"svc_{cid}")
        except Exception: pass

    def _ask_service(self, cid):
        try:
            self.get_d(cid)['pending_review'] = 'service'
            self.bot.send_message(cid, "Как вам наш сервис? Поделитесь мнением о работе Джарвиса 🙏", reply_markup=types.ReplyKeyboardRemove())
        except Exception as e: log.error(f"_ask_service: {e}")

    def _ask_product(self, cid):
        try:
            self.get_d(cid)['pending_review'] = 'product'
            self.bot.send_message(cid, "Прошли сутки. Вы получили заказ? Оставьте отзыв! ⭐", reply_markup=types.ReplyKeyboardRemove())
        except Exception as e: log.error(f"_ask_product: {e}")


class ClientHandler(BaseHandler):
    def __init__(self, bot, user_data, scheduler, admin_ids, reviews: ReviewHandler):
        super().__init__(bot, user_data, scheduler, admin_ids)
        self.reviews = reviews

    def register(self):
        @self.bot.message_handler(func=lambda m: self.user_data.get(m.chat.id, {}).get('step') == 'city' and not self.user_data.get(m.chat.id, {}).get('admin_step'))
        def city(m): self._city(m)

        @self.bot.message_handler(func=lambda m: self.user_data.get(m.chat.id, {}).get('step') == 'brand' and not self.user_data.get(m.chat.id, {}).get('admin_step'))
        def brand(m): self._brand(m)

        @self.bot.message_handler(func=lambda m: self.user_data.get(m.chat.id, {}).get('step') == 'cat' and not self.user_data.get(m.chat.id, {}).get('admin_step'))
        def cat(m): self._cat(m)

        @self.bot.message_handler(func=lambda m: self.user_data.get(m.chat.id, {}).get('step') == 'price' and not self.user_data.get(m.chat.id, {}).get('admin_step'))
        def price(m): self._price(m)

        @self.bot.message_handler(func=lambda m: self.user_data.get(m.chat.id, {}).get('step') == 'size' and not self.user_data.get(m.chat.id, {}).get('admin_step'))
        def size(m): self._size(m)

        @self.bot.message_handler(func=lambda m: m.text and "Я написал продавцу" in m.text and self.user_data.get(m.chat.id, {}).get('step') == 'done')
        def confirm(m): self._confirm(m)

        @self.bot.message_handler(func=lambda m: m.text and "Уведомить меня" in m.text)
        def waiting(m): self._waiting(m)

        @self.bot.message_handler(func=lambda m: m.text and "Оставить заявку" in m.text)
        def request(m): self._request(m)

    def _city(self, m):
        cid = m.chat.id
        try:
            db.save_user(m.from_user.id, username=m.from_user.username, first_name=m.from_user.first_name)
            cities = db.get_unique_cities()
            chosen = m.text.strip()
            cl = [c.lower() for c in cities]
            if chosen.lower() not in cl:
                self.bot.send_message(cid, "Пожалуйста, выберите город из списка:", reply_markup=make_reply_keyboard(cities))
                return
            real = cities[cl.index(chosen.lower())]
            self.user_data[cid]['city'] = real
            self.user_data[cid]['step'] = 'brand'
            brands = db.get_unique_brands(real)
            if not brands:
                self.bot.send_message(cid, f"😔 В городе *{real}* товаров пока нет.", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
                return
            self.bot.send_message(cid, f"📍 Город: *{real}*\n\nВыберите бренд:", parse_mode="Markdown", reply_markup=make_reply_keyboard(brands))
        except Exception as e: log.error(f"_city: {e}")

    def _brand(self, m):
        cid = m.chat.id
        try:
            d = self.get_d(cid)
            city = d.get('city', '')
            brands = db.get_unique_brands(city)
            bl = [b.lower() for b in brands]
            chosen = m.text.strip()
            if chosen.lower() not in bl:
                self.bot.send_message(cid, "Пожалуйста, выберите бренд из списка:", reply_markup=make_reply_keyboard(brands))
                return
            real = brands[bl.index(chosen.lower())]
            d['brand'] = real
            d['step'] = 'cat'
            cats = db.get_unique_categories(city, real)
            if not cats:
                self.bot.send_message(cid, f"😔 Для бренда *{real}* категорий не найдено.", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
                return
            self.bot.send_message(cid, f"👟 Бренд: *{real}*\n\nВыберите категорию:", parse_mode="Markdown", reply_markup=make_reply_keyboard(cats))
        except Exception as e: log.error(f"_brand: {e}")

    def _cat(self, m):
        cid = m.chat.id
        try:
            d = self.get_d(cid)
            city = d.get('city', '')
            brand = d.get('brand', '')
            cats = db.get_unique_categories(city, brand)
            cl = [c.lower() for c in cats]
            chosen = m.text.strip()
            if chosen.lower() not in cl:
                self.bot.send_message(cid, "Пожалуйста, выберите категорию из списка:", reply_markup=make_reply_keyboard(cats))
                return
            real = cats[cl.index(chosen.lower())]
            d['cat'] = real
            d['step'] = 'price'
            pm = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
            pm.add(types.KeyboardButton("💰 Любая цена"))
            self.bot.send_message(cid, f"📦 Категория: *{real}*\n\nВведите ваш *бюджет* (KZT) или нажмите «Любая цена»:", parse_mode="Markdown", reply_markup=pm)
        except Exception as e: log.error(f"_cat: {e}")

    def _price(self, m):
        cid = m.chat.id
        try:
            text = m.text.strip()
            price = 999_999_999 if "Любая цена" in text else int(''.join(filter(str.isdigit, text)) or '999999999')
            self.user_data[cid].update({'price': price, 'step': 'size'})
            self.bot.send_message(cid, "📏 Введите ваш *размер*:", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
        except Exception as e: log.error(f"_price: {e}")

    def _size(self, m):
        cid = m.chat.id
        try:
            d = self.get_d(cid)
            d['size'] = m.text.strip()
            d['step'] = 'done'
            brand = d.get('brand', '')
            cat = d.get('cat', '')
            city = d.get('city', '')
            price = d.get('price', 999_999_999)
            size = d['size']
            uid = m.from_user.id
            query_text = f"{brand} {cat}"

            res = db.search_items_v4(brand, cat, city, price, size=size)

            if res:
                item = res[0]
                (item_id, i_brand, i_cat, i_price, wa_link,
                 seller_name, material, color, i_city,
                 description, photo_url, sizes, partner_id) = item

                db.log_action(uid, city, size, query_text, "Описание выдано", seller=seller_name)

                caption = (
                    f"🏷 *Джарвис нашёл вариант:*\n\n"
                    f"👟 *Бренд:* {i_brand}\n"
                    f"📦 *Категория:* {i_cat}\n"
                    f"💰 *Цена:* {i_price} KZT\n"
                    f"🧵 *Материал:* {material}\n"
                    f"🎨 *Цвет:* {color}\n"
                    f"📍 *Город:* {i_city}\n"
                    f"📏 *Размеры:* {sizes or 'уточните у продавца'}\n\n"
                    f"📝 *Описание:*\n{description}"
                )

                wa_url = build_wa_url(wa_link, i_brand, i_cat, material, size, partner_id)
                d.update({'wa_url': wa_url, 'seller_name': seller_name, 'item_id': item_id, 'material': material})

                om = types.InlineKeyboardMarkup()
                om.add(types.InlineKeyboardButton("📲 Оформить заказ", url=wa_url))
                cm = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                cm.add(types.KeyboardButton("✅ Я написал продавцу"))

                if photo_url and photo_url.strip():
                    try:
                        self.bot.send_photo(cid, photo=photo_url.strip(), caption=caption, parse_mode="Markdown", reply_markup=om)
                    except Exception:
                        self.bot.send_message(cid, caption, parse_mode="Markdown", reply_markup=om)
                else:
                    self.bot.send_message(cid, caption, parse_mode="Markdown", reply_markup=om)

                self.bot.send_message(cid, "После того как написали продавцу — нажмите кнопку ниже:", reply_markup=cm)
                self.reviews.schedule_service(cid)

            else:
                db.log_search_fail(uid, query_text, size)
                similar = db.search_similar(cat, city)

                if similar:
                    s = similar[0]
                    (s_id, s_brand, s_cat, s_price, s_wa, s_seller, s_material, s_color, s_city, s_desc, s_photo, s_sizes, s_pid) = s
                    s_caption = (
                        f"😔 *Точного совпадения нет.*\n\n"
                        f"🔎 *Но нашёл похожий вариант:*\n\n"
                        f"👟 *Бренд:* {s_brand}\n"
                        f"📦 *Категория:* {s_cat}\n"
                        f"💰 *Цена:* {s_price} KZT\n"
                        f"🧵 *Материал:* {s_material}\n"
                        f"🎨 *Цвет:* {s_color}\n"
                        f"📍 *Город:* {s_city}\n\n"
                        f"📝 *Описание:*\n{s_desc}"
                    )
                    s_wa_url = build_wa_url(s_wa, s_brand, s_cat, s_material, size, s_pid)
                    d.update({'wa_url': s_wa_url, 'seller_name': s_seller, 'step': 'done'})
                    sm = types.InlineKeyboardMarkup()
                    sm.add(types.InlineKeyboardButton("📲 Оформить этот вариант", url=s_wa_url))
                    rm = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                    rm.row(types.KeyboardButton("✅ Я написал продавцу"))
                    rm.row(types.KeyboardButton("📋 Оставить заявку на мой вариант"))
                    if s_photo and s_photo.strip():
                        try:
                            self.bot.send_photo(cid, photo=s_photo.strip(), caption=s_caption, parse_mode="Markdown", reply_markup=sm)
                        except Exception:
                            self.bot.send_message(cid, s_caption, parse_mode="Markdown", reply_markup=sm)
                    else:
                        self.bot.send_message(cid, s_caption, parse_mode="Markdown", reply_markup=sm)
                    self.bot.send_message(cid, "Хотите этот вариант или оставить заявку на свой?", reply_markup=rm)
                else:
                    nm = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                    nm.add(types.KeyboardButton("📋 Оставить заявку"))
                    nm.add(types.KeyboardButton("🔔 Уведомить меня когда появится"))
                    self.bot.send_message(cid, "😔 *Товар не найден.*\n\nМожете оставить заявку или подписаться на уведомление!", parse_mode="Markdown", reply_markup=nm)

                self.notify_admin(
                    f"🔔 *Товар не найден!*\n\n"
                    f"👤 Пользователь: `{uid}`\n"
                    f"🔍 Запрос: {brand} | {cat}\n"
                    f"📍 Город: {city}\n"
                    f"📏 Размер: {size}\n"
                    f"💰 Бюджет: {price if price < 999_999_999 else 'любой'} KZT"
                )
        except Exception as e:
            log.error(f"_size: {e}")
            self.bot.send_message(cid, f"❌ Ошибка при поиске: {e}")

    def _confirm(self, m):
        cid = m.chat.id
        try:
            d = self.get_d(cid)
            uid = m.from_user.id
            seller = d.get('seller_name', '')
            brand = d.get('brand', '')
            cat = d.get('cat', '')
            city = d.get('city', '')
            size = d.get('size', '')
            query_text = f"{brand} {cat}"
            self.reviews.cancel_service(cid)
            db.log_action(uid, city, size, query_text, "Перешел на ватсап", seller=seller)
            db.log_action(uid, city, size, query_text, "whatsapp_redirect", seller=seller)
            self.bot.send_message(cid, "🎉 *Отлично! Ваш заказ зафиксирован.*\n\nМенеджер свяжется с вами в ближайшее время!", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
            self.reviews.schedule_product(cid)
        except Exception as e: log.error(f"_confirm: {e}")

    def _waiting(self, m):
        cid = m.chat.id
        d = self.get_d(cid)
        db.add_to_waiting_list(m.from_user.id, d.get('cat', '—'), d.get('size', '—'), d.get('city', '—'), d.get('brand', '—'))
        self.bot.send_message(cid, f"🔔 *Подписка оформлена!*\n\nКак только появится *{d.get('brand')} {d.get('cat')}* в {d.get('city')} (р.{d.get('size')}) — напишем!", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
        d['step'] = None

    def _request(self, m):
        cid = m.chat.id
        d = self.get_d(cid)
        uid = m.from_user.id
        price = d.get('price', 0)
        self.notify_admin(
            f"📋 *НОВАЯ ЗАЯВКА*\n\n"
            f"👤 Пользователь: `{uid}`\n"
            f"📱 Username: @{m.from_user.username or 'нет'}\n"
            f"👋 Имя: {m.from_user.first_name or 'нет'}\n\n"
            f"🔍 *Что ищет:*\n"
            f"👟 Бренд: {d.get('brand', '—')}\n"
            f"📦 Категория: {d.get('cat', '—')}\n"
            f"📍 Город: {d.get('city', '—')}\n"
            f"📏 Размер: {d.get('size', '—')}\n"
            f"💰 Бюджет: {price if price < 999_999_999 else 'любой'} KZT"
        )
        self.bot.send_message(cid, "✅ *Заявка отправлена!*\n\nМенеджер свяжется с вами в течение 48 часов.", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
        d['step'] = None


class AdminHandler(BaseHandler):
    def register(self):
        @self.bot.message_handler(func=lambda m: self.is_admin(m.from_user.id) and self.user_data.get(m.chat.id, {}).get('step') == 'admin_menu')
        def menu(m): self._menu(m)

        @self.bot.message_handler(func=lambda m: self.is_admin(m.from_user.id) and self.user_data.get(m.chat.id, {}).get('admin_step') in STEP_TO_FIELD)
        def add_steps(m): self._add_item_steps(m)

        @self.bot.message_handler(func=lambda m: self.is_admin(m.from_user.id) and self.user_data.get(m.chat.id, {}).get('step') == 'item_manage')
        def item_manage(m): self._item_manage(m)

        @self.bot.message_handler(func=lambda m: self.is_admin(m.from_user.id) and self.user_data.get(m.chat.id, {}).get('step') == 'wait_item_id')
        def wait_id(m): self._wait_item_id(m)

        @self.bot.message_handler(func=lambda m: self.is_admin(m.from_user.id) and self.user_data.get(m.chat.id, {}).get('step') == 'edit_field')
        def edit_field(m): self._edit_field(m)

        @self.bot.message_handler(func=lambda m: self.is_admin(m.from_user.id) and self.user_data.get(m.chat.id, {}).get('step') == 'edit_value')
        def edit_value(m): self._edit_value(m)

        @self.bot.message_handler(func=lambda m: self.is_admin(m.from_user.id) and self.user_data.get(m.chat.id, {}).get('step') == 'broadcast')
        def broadcast(m): self._broadcast(m)

        @self.bot.message_handler(func=lambda m: self.is_admin(m.from_user.id) and self.user_data.get(m.chat.id, {}).get('step') == 'stats_detail')
        def stats_detail(m): self._stats_detail(m)

    def _menu(self, m):
        cid = m.chat.id
        text = m.text.strip()
        try:
            if "Добавить товар" in text:
                self.user_data[cid]['admin_step'] = 'ai_brand'
                self.user_data[cid]['new_item'] = {}
                self.bot.send_message(cid, "➕ *Добавление нового товара*\n\nШаг 1/12 — Введите *бренд*:", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())

            elif "Список товаров" in text:
                items = db.get_all_items()
                if not items:
                    self.bot.send_message(cid, "В базе нет товаров.", reply_markup=admin_menu_keyboard())
                    return
                for i in range(0, len(items), 10):
                    chunk = items[i:i+10]
                    lines = [f"📋 *Список ({i+1}-{i+len(chunk)} из {len(items)}):*\n"]
                    for row in chunk:
                        iid, brand, category, price, city, photo_url, sizes, pid = row
                        photo = "🖼" if photo_url else "—"
                        sz = sizes or "—"
                        p_tag = f" [П{pid}]" if pid else ""
                        lines.append(f"ID `{iid}` | {brand} | {category} | {price} KZT | {city} | р.{sz}{p_tag} {photo}")
                    self.bot.send_message(cid, "\n".join(lines), parse_mode="Markdown")
                mgmt = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                mgmt.row(types.KeyboardButton("✏️ Редактировать товар"))
                mgmt.row(types.KeyboardButton("🗑 Удалить товар"))
                mgmt.row(types.KeyboardButton("🔙 Назад в меню"))
                self.bot.send_message(cid, "Выберите действие:", reply_markup=mgmt)
                self.user_data[cid]['step'] = 'item_manage'

            elif "Статистика" in text:
                rows = db.get_logs_stats()
                fails = db.get_search_fails_count()
                total = db.get_users_count()
                lines = [f"📊 *Статистика:*\n\n👥 Всего пользователей: *{total}*\n"]
                for action_type, cnt in rows:
                    lines.append(f"• {action_type}: *{cnt}*")
                lines.append(f"\n🔍 Запросов без результата: *{fails}*")
                sm = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                sm.row(types.KeyboardButton("📋 Результаты поиска"))
                sm.row(types.KeyboardButton("🔍 Неудачные поиски"))
                sm.row(types.KeyboardButton("🔙 Назад в меню"))
                self.bot.send_message(cid, "\n".join(lines), parse_mode="Markdown", reply_markup=sm)
                self.user_data[cid]['step'] = 'stats_detail'

            elif "Отзывы о сервисе" in text:
                rows = db.get_recent_reviews_service()
                if not rows:
                    self.bot.send_message(cid, "Отзывов о сервисе пока нет.", reply_markup=admin_menu_keyboard())
                    return
                lines = ["⭐ *Отзывы о сервисе:*\n"]
                for uid2, review, ts in rows:
                    dt = ts.strftime("%d.%m %H:%M") if ts else "—"
                    lines.append(f"`{dt}` | {uid2}\n{review}\n")
                self.bot.send_message(cid, "\n".join(lines)[:4000], parse_mode="Markdown", reply_markup=admin_menu_keyboard())

            elif "Отзывы о товарах" in text:
                rows = db.get_recent_reviews_product()
                if not rows:
                    self.bot.send_message(cid, "Отзывов о товарах пока нет.", reply_markup=admin_menu_keyboard())
                    return
                lines = ["📦 *Отзывы о товарах:*\n"]
                for uid2, review, ts in rows:
                    dt = ts.strftime("%d.%m %H:%M") if ts else "—"
                    lines.append(f"`{dt}` | {uid2}\n{review}\n")
                self.bot.send_message(cid, "\n".join(lines)[:4000], parse_mode="Markdown", reply_markup=admin_menu_keyboard())

            elif "Рассылка" in text:
                total = db.get_users_count()
                self.user_data[cid]['step'] = 'broadcast'
                cm = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                cm.add(types.KeyboardButton("❌ Отмена"))
                self.bot.send_message(cid, f"📣 *Рассылка*\n\n👥 Сообщение получат *{total}* пользователей.\n\nНапишите текст:", parse_mode="Markdown", reply_markup=cm)

            elif "Выгрузить" in text:
                self.bot.send_message(cid, "📥 Выгружаю данные...", reply_markup=types.ReplyKeyboardRemove())
                self._export_excel(cid)

            else:
                self.bot.send_message(cid, "Выберите действие из меню:", reply_markup=admin_menu_keyboard())

        except Exception as e:
            log.error(f"AdminHandler._menu: {e}")
            self.bot.send_message(cid, f"❌ Ошибка: {e}", reply_markup=admin_menu_keyboard())

    def _add_item_steps(self, m):
        cid = m.chat.id
        d = self.get_d(cid)
        current_step = d['admin_step']
        field = STEP_TO_FIELD[current_step]
        value = m.text.strip()
        try:
            if field == 'photo_url' and value.lower() in ('нет', 'no', '-', '—'):
                value = None
            elif field == 'photo_url' and value and not any(value.lower().endswith(x) for x in ('.jpg','.jpeg','.png','.webp','.gif')):
                self.bot.send_message(cid, "Внимание! Ссылка не выглядит как прямое фото. Сохраняю как есть...")
            elif field == 'sizes' and value.lower() in ('нет', 'no', '—'):
                value = None
            elif field == 'sizes' and value:
                mr = re.match(r'^(\d+)\s*[-–]\s*(\d+)$', value.strip())
                if mr:
                    value = ';'.join(str(s) for s in range(int(mr.group(1)), int(mr.group(2))+1))
                else:
                    value = ';'.join([s.strip() for s in value.replace(',', ';').replace(' ', ';').split(';') if s.strip()])
            elif field == 'partner_id':
                if value.lower() in ('нет', 'no', '-', '—', '0'):
                    value = None
                else:
                    try:
                        pid = int(value)
                        partner = db.get_partner_by_id(pid)
                        if partner:
                            value = pid
                            self.bot.send_message(cid, f"✅ Партнёр: *{partner[1]}*", parse_mode="Markdown")
                        else:
                            self.bot.send_message(cid, f"❌ Партнёр ID {pid} не найден. Сохраняю без партнёра.")
                            value = None
                    except ValueError:
                        value = None

            if field not in ('photo_url', 'sizes', 'partner_id') and value and ';' in str(value):
                parts = [p.strip() for p in value.split(';') if p.strip()]
                if len(parts) > 1:
                    d['new_item'][field] = parts[0]
                    d['bulk_values'] = {field: parts}
                    self.bot.send_message(cid, f"📦 Обнаружено {len(parts)} значений: {', '.join(parts)}\nБуду сохранять {len(parts)} отдельных записей!")
                else:
                    d['new_item'][field] = value
            else:
                d['new_item'][field] = value

            if current_step == 'ai_sizes':
                partners = db.get_all_partners()
                if partners:
                    lines = ["Шаг 12/12 — *Выберите партнёра* (введите ID) или напишите «нет»:\n"]
                    for p in partners:
                        status = "✅" if p[3] else "❌"
                        lines.append(f"{status} ID `{p[0]}` — {p[1]} | {p[2] or '—'}")
                    self.bot.send_message(cid, "\n".join(lines), parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
                else:
                    self.bot.send_message(cid, "Шаг 12/12 — Партнёров нет. Напишите «нет»:")
                d['admin_step'] = 'ai_partner'
                return

            next_step = next_prompt = None
            for cs, nxt, prompt in ADMIN_ITEM_STEPS:
                if cs == current_step:
                    next_step, next_prompt = nxt, prompt
                    break

            if next_step and current_step != 'ai_sizes':
                d['admin_step'] = next_step
                self.bot.send_message(cid, next_prompt, parse_mode="Markdown")
            else:
                self._save_item(cid, d)

        except Exception as e:
            log.error(f"AdminHandler._add_item_steps: {e}")
            self.bot.send_message(cid, f"❌ Ошибка: {e}")

    def _save_item(self, cid, d):
        item = d['new_item']
        price = int(''.join(filter(str.isdigit, str(item.get('price', '0')))) or '0')
        bulk = d.get('bulk_values', {})
        if bulk:
            bulk_field = list(bulk.keys())[0]
            added_ids = []
            for bval in bulk[bulk_field]:
                ic = dict(item)
                ic[bulk_field] = bval
                p = int(''.join(filter(str.isdigit, str(ic.get('price', '0')))) or '0')
                bid = db.add_item(brand=ic['brand'], category=ic['category'], price=p,
                    material=ic['material'], color=ic['color'], city=ic['city'],
                    wa_link=ic['wa_link'], description=ic['description'],
                    seller_name=ic['seller_name'], photo_url=ic.get('photo_url'),
                    sizes=ic.get('sizes'), partner_id=ic.get('partner_id'))
                added_ids.append(str(bid))
            self.bot.send_message(cid, f"✅ Добавлено {len(added_ids)} товаров!\nID: {', '.join(added_ids)}")
            d['bulk_values'] = {}
        else:
            new_id = db.add_item(brand=item['brand'], category=item['category'], price=price,
                material=item['material'], color=item['color'], city=item['city'],
                wa_link=item['wa_link'], description=item['description'],
                seller_name=item['seller_name'], photo_url=item.get('photo_url'),
                sizes=item.get('sizes'), partner_id=item.get('partner_id'))
            pid_info = f"Партнёр ID: {item.get('partner_id')}" if item.get('partner_id') else "без партнёра"
            self.bot.send_message(cid,
                f"✅ Товар добавлен! ID: {new_id}\n\n"
                f"Бренд: {item['brand']}\nКатегория: {item['category']}\nЦена: {price} KZT\n"
                f"Материал: {item['material']}\nЦвет: {item['color']}\nГород: {item['city']}\n"
                f"WA: {item['wa_link']}\nПродавец: {item['seller_name']}\n"
                f"Фото: {item.get('photo_url') or 'не указано'}\n"
                f"Размеры: {item.get('sizes') or 'не указаны'}\n{pid_info}")
        d['admin_step'] = None
        d['step'] = 'admin_menu'
        self.bot.send_message(cid, "Вернуться в меню:", reply_markup=admin_menu_keyboard())

    def _item_manage(self, m):
        cid = m.chat.id
        text = m.text.strip()
        if "Редактировать" in text:
            self.user_data[cid].update({'step': 'wait_item_id', 'action': 'edit'})
            self.bot.send_message(cid, "Введите *ID товара* для редактирования:", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
        elif "Удалить" in text:
            self.user_data[cid].update({'step': 'wait_item_id', 'action': 'delete'})
            self.bot.send_message(cid, "Введите *ID товара* для удаления:", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())
        else:
            self.user_data[cid]['step'] = 'admin_menu'
            self.bot.send_message(cid, "Вернулись в меню.", reply_markup=admin_menu_keyboard())

    def _wait_item_id(self, m):
        cid = m.chat.id
        action = self.user_data[cid].get('action', 'edit')
        try:
            iid = int(m.text.strip())
            item = db.get_item_by_id(iid)
            if not item:
                self.bot.send_message(cid, f"❌ Товар с ID {iid} не найден. Попробуйте ещё раз:")
                return
            def esc(v): return str(v or '').replace('_', '\\_').replace('*', '\\*').replace('`', '\\`')
            (iid2, brand, cat, price, material, color, city, wa_link, desc, seller, photo, sizes, pid) = item
            info = (
                f"📦 *Товар ID {iid2}:*\n"
                f"👟 Бренд: {esc(brand)}\n📦 Категория: {esc(cat)}\n💰 Цена: {price} KZT\n"
                f"🧵 Материал: {esc(material)}\n🎨 Цвет: {esc(color)}\n📍 Город: {esc(city)}\n"
                f"🔗 WA: {esc(wa_link)}\n👤 Продавец: {esc(seller)}\n"
                f"📏 Размеры: {esc(sizes)}\n🤝 Партнёр ID: {pid or '—'}"
            )
            self.bot.send_message(cid, info, parse_mode="Markdown")
            if action == 'delete':
                db.delete_item(iid)
                self.bot.send_message(cid, f"🗑 Товар ID {iid} удалён!", reply_markup=admin_menu_keyboard())
                self.user_data[cid]['step'] = 'admin_menu'
            else:
                self.user_data[cid].update({'edit_item_id': iid, 'step': 'edit_field'})
                fm = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
                for f in ['brand', 'category', 'price', 'material', 'color', 'city', 'wa_link', 'description', 'seller_name', 'photo_url', 'sizes', 'partner_id']:
                    fm.add(types.KeyboardButton(f))
                self.bot.send_message(cid, "Выберите поле для редактирования:", reply_markup=fm)
        except ValueError:
            self.bot.send_message(cid, "❌ Введите корректный числовой ID:")

    def _edit_field(self, m):
        cid = m.chat.id
        field = m.text.strip()
        allowed = {'brand', 'category', 'price', 'material', 'color', 'city', 'wa_link', 'description', 'seller_name', 'photo_url', 'sizes', 'partner_id'}
        if field not in allowed:
            self.bot.send_message(cid, "❌ Недопустимое поле.")
            return
        self.user_data[cid].update({'edit_field': field, 'step': 'edit_value'})
        self.bot.send_message(cid, f"Введите новое значение для *{field}*:", parse_mode="Markdown", reply_markup=types.ReplyKeyboardRemove())

    def _edit_value(self, m):
        cid = m.chat.id
        field = self.user_data[cid].get('edit_field')
        iid = self.user_data[cid].get('edit_item_id')
        value = m.text.strip()
        try:
            if field == 'price':
                value = int(''.join(filter(str.isdigit, value)) or '0')
            elif field == 'partner_id':
                value = int(value) if value.isdigit() else None
            db.update_item_field(iid, field, value)
            self.bot.send_message(cid, f"✅ Поле *{field}* обновлено!", parse_mode="Markdown", reply_markup=admin_menu_keyboard())
            self.user_data[cid]['step'] = 'admin_menu'
        except Exception as e:
            self.bot.send_message(cid, f"❌ Ошибка: {e}")

    def _broadcast(self, m):
        cid = m.chat.id
        text = m.text.strip()
        if "Отмена" in text:
            self.user_data[cid]['step'] = 'admin_menu'
            self.bot.send_message(cid, "Рассылка отменена.", reply_markup=admin_menu_keyboard())
            return
        user_ids = db.get_all_user_ids()
        sent = failed = 0
        for uid in user_ids:
            try:
                self.bot.send_message(uid, text)
                sent += 1
            except Exception:
                failed += 1
        self.bot.send_message(cid, f"📣 Рассылка завершена!\n✅ Отправлено: {sent}\n❌ Не доставлено: {failed}", reply_markup=admin_menu_keyboard())
        self.user_data[cid]['step'] = 'admin_menu'

    def _stats_detail(self, m):
        cid = m.chat.id
        text = m.text.strip()
        try:
            if "Результаты поиска" in text:
                rows = db.repo.logs.get_traffic_for_export()
                if not rows:
                    self.bot.send_message(cid, "Пока нет данных.", reply_markup=admin_menu_keyboard())
                else:
                    lines = ["📋 *Последние результаты поиска:*\n"]
                    for row in rows[:30]:
                        dt = row[7].strftime("%d.%m %H:%M") if row[7] else "—"
                        lines.append(f"`{dt}` | {row[3]} | {row[4]} | р.{row[5]}")
                    self.bot.send_message(cid, "\n".join(lines), parse_mode="Markdown", reply_markup=admin_menu_keyboard())
            elif "Неудачные поиски" in text:
                with db.repo.logs.db.get_connection() as conn:
                    with conn.cursor() as cur:
                        cur.execute("SELECT query_text, requested_size, timestamp FROM logs_search_requests ORDER BY timestamp DESC LIMIT 30")
                        rows = cur.fetchall()
                if not rows:
                    self.bot.send_message(cid, "Неудачных поисков нет. 🎉", reply_markup=admin_menu_keyboard())
                else:
                    lines = ["🔍 *Неудачные поиски:*\n"]
                    for query, size, ts in rows:
                        dt = ts.strftime("%d.%m %H:%M") if ts else "—"
                        lines.append(f"`{dt}` | {query} | р.{size}")
                    self.bot.send_message(cid, "\n".join(lines), parse_mode="Markdown", reply_markup=admin_menu_keyboard())
            else:
                self.user_data[cid]['step'] = 'admin_menu'
                self.bot.send_message(cid, "Вернулись в меню.", reply_markup=admin_menu_keyboard())
            self.user_data[cid]['step'] = 'admin_menu'
        except Exception as e:
            log.error(f"_stats_detail: {e}")
            self.bot.send_message(cid, f"❌ Ошибка: {e}", reply_markup=admin_menu_keyboard())

    def _export_excel(self, cid):
        try:
            from openpyxl.styles import Font, PatternFill
            traffic_rows = db.repo.logs.get_traffic_for_export()
            sales_rows = db.repo.logs.get_sales_for_export()
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_t = pd.DataFrame(traffic_rows, columns=['user_id', 'username', 'имя', 'запрос', 'город', 'размер', 'действие', 'время'])
                df_t['username'] = df_t['username'].apply(lambda x: f"@{x}" if x and x != '—' else '—')
                df_t.to_excel(writer, sheet_name='Traffic', index=False)
                ws1 = writer.sheets['Traffic']
                for cell in ws1[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(fill_type='solid', fgColor='2E75B6')
                for col in ws1.columns:
                    ws1.column_dimensions[col[0].column_letter].width = 20

                df_s = pd.DataFrame(sales_rows, columns=['user_id', 'username', 'имя', 'товар', 'город', 'размер', 'продавец', 'партнёр', 'контакт_партнёра', 'время'])
                df_s['username'] = df_s['username'].apply(lambda x: f"@{x}" if x and x != '—' else '—')
                df_s.to_excel(writer, sheet_name='Sales', index=False)
                ws2 = writer.sheets['Sales']
                for cell in ws2[1]:
                    cell.font = Font(bold=True, color='FFFFFF')
                    cell.fill = PatternFill(fill_type='solid', fgColor='375623')
                for col in ws2.columns:
                    ws2.column_dimensions[col[0].column_letter].width = 20

            output.seek(0)
            now = datetime.now().strftime("%d-%m-%Y_%H-%M")
            self.bot.send_document(cid, document=output, visible_file_name=f"Jarvis_Report_{now}.xlsx",
                caption=f"📊 *Отчёт Jarvis*\n📅 {now}\n\n📋 Traffic: {len(traffic_rows)}\n💰 Sales: {len(sales_rows)}",
                parse_mode="Markdown")
            self.bot.send_message(cid, "✅ Excel отчёт готов!", reply_markup=admin_menu_keyboard())
        except Exception as e:
            log.error(f"_export_excel: {e}")
            self.bot.send_message(cid, f"❌ Ошибка: {e}", reply_markup=admin_menu_keyboard())


class VendorHandler(BaseHandler):
    def register(self):
        @self.bot.message_handler(func=lambda m: self.user_data.get(m.chat.id, {}).get('step') == 'vendor_menu')
        def menu(m): self._menu(m)

    def _menu(self, m):
        cid = m.chat.id
        text = m.text.strip()
        try:
            partner_id = self.user_data[cid].get('vendor_partner_id')
            if not partner_id:
                self.bot.send_message(cid, "Сессия истекла. Используйте /vendor.")
                return
            if "Добавить товар" in text:
                self.user_data[cid].update({'step': 'admin_menu', 'admin_step': 'ai_brand', 'new_item': {}})
            elif "Мои товары" in text:
                items = db.get_partner_items(partner_id)
                if not items:
                    self.bot.send_message(cid, "У вас пока нет товаров.", reply_markup=vendor_menu_keyboard())
                    return
                lines = [f"📦 *Ваши товары ({len(items)}):*\n"]
                for iid, brand, category, price, city, sizes, photo_url in items:
                    lines.append(f"ID `{iid}` | {brand} | {category} | {price} KZT | р.{sizes or '—'}")
                self.bot.send_message(cid, "\n".join(lines), parse_mode="Markdown", reply_markup=vendor_menu_keyboard())
            elif "Статистика" in text:
                stats = db.get_partner_stats(partner_id)
                self.bot.send_message(cid,
                    f"📊 *Ваша статистика:*\n\n📦 Товаров: *{stats['items']}*\n👁 Просмотров: *{stats['views']}*\n📲 Переходов на WA: *{stats['wa_clicks']}*",
                    parse_mode="Markdown", reply_markup=vendor_menu_keyboard())
            else:
                self.bot.send_message(cid, "Выберите действие:", reply_markup=vendor_menu_keyboard())
        except Exception as e:
            log.error(f"VendorHandler._menu: {e}")
            self.bot.send_message(cid, f"❌ Ошибка: {e}", reply_markup=vendor_menu_keyboard())


class JarvisBot:
    """Главный класс бота. Собирает все хендлеры и запускает polling."""

    ADMIN_IDS = [5720010749]

    def __init__(self):
        self.bot = telebot.TeleBot(TOKEN, parse_mode=None)
        self.scheduler = BackgroundScheduler()
        self.user_data = {}

        db.init_db()

        self.reviews = ReviewHandler(self.bot, self.user_data, self.scheduler, self.ADMIN_IDS)
        self.client  = ClientHandler(self.bot, self.user_data, self.scheduler, self.ADMIN_IDS, self.reviews)
        self.admin   = AdminHandler(self.bot, self.user_data, self.scheduler, self.ADMIN_IDS)
        self.vendor  = VendorHandler(self.bot, self.user_data, self.scheduler, self.ADMIN_IDS)

    def _setup_commands(self):
        @self.bot.message_handler(commands=['start'])
        def cmd_start(m):
            try:
                db.save_user(m.from_user.id, username=m.from_user.username, first_name=m.from_user.first_name)
                self.user_data[m.chat.id] = {'step': 'city', 'pending_review': None, 'admin_step': None, 'new_item': {}}
                cities = db.get_unique_cities()
                if not cities:
                    self.bot.send_message(m.chat.id, "😔 В базе пока нет товаров. Обратитесь к менеджеру.", reply_markup=types.ReplyKeyboardRemove())
                    return
                self.bot.send_message(m.chat.id, "👋 *Джарвис на связи!*\n\nВыберите ваш город:", parse_mode="Markdown", reply_markup=make_reply_keyboard(cities))
            except Exception as e:
                log.error(f"cmd_start: {e}")
                self.bot.send_message(m.chat.id, f"❌ Ошибка: {e}")

        @self.bot.message_handler(commands=['admin'])
        def cmd_admin(m):
            try:
                if m.from_user.id not in self.ADMIN_IDS:
                    self.bot.send_message(m.chat.id, "⛔ У вас нет доступа.")
                    return
                self.user_data[m.chat.id] = {'step': 'admin_menu', 'pending_review': None, 'admin_step': None, 'new_item': {}}
                total = db.get_users_count()
                self.bot.send_message(m.chat.id, f"🛠 *Админ-панель Джарвис*\n👥 Пользователей в базе: *{total}*\n\nВыберите действие:", parse_mode="Markdown", reply_markup=admin_menu_keyboard())
            except Exception as e: log.error(f"cmd_admin: {e}")

        @self.bot.message_handler(commands=['partners'])
        def cmd_partners(m):
            try:
                if m.from_user.id not in self.ADMIN_IDS:
                    self.bot.send_message(m.chat.id, "⛔ Нет доступа.")
                    return
                partners = db.get_all_partners(only_active=False)
                if not partners:
                    self.bot.send_message(m.chat.id, "Партнёров пока нет.\n\nДобавить: /add_partner Название | Контакт")
                    return
                lines = ["👥 *Список партнёров:*\n"]
                for p in partners:
                    status = "✅" if p[3] else "❌"
                    lines.append(f"{status} ID `{p[0]}` — *{p[1]}*\n   📞 {p[2] or '—'}")
                lines.append("\nДобавить: /add_partner Название | Контакт")
                lines.append("Деактивировать: /deactivate_partner [id]")
                self.bot.send_message(m.chat.id, "\n".join(lines), parse_mode="Markdown")
            except Exception as e: log.error(f"cmd_partners: {e}")

        @self.bot.message_handler(commands=['add_partner'])
        def cmd_add_partner(m):
            try:
                if m.from_user.id not in self.ADMIN_IDS:
                    self.bot.send_message(m.chat.id, "⛔ Нет доступа.")
                    return
                parts = m.text.split(None, 1)
                if len(parts) < 2:
                    self.bot.send_message(m.chat.id, "Использование: /add_partner Название | Контакт\nПример: /add_partner Магазин | +77001234567")
                    return
                data = parts[1].split("|", 1)
                name = data[0].strip()
                contact = data[1].strip() if len(data) > 1 else None
                new_id = db.add_partner(name, contact)
                self.bot.send_message(m.chat.id, f"✅ *Партнёр добавлен!*\n\n🆔 ID: `{new_id}`\n👤 Название: {name}\n📞 Контакт: {contact or '—'}", parse_mode="Markdown")
            except Exception as e:
                log.error(f"cmd_add_partner: {e}")
                self.bot.send_message(m.chat.id, f"❌ Ошибка: {e}")

        @self.bot.message_handler(commands=['deactivate_partner'])
        def cmd_deactivate(m):
            try:
                if m.from_user.id not in self.ADMIN_IDS:
                    self.bot.send_message(m.chat.id, "⛔ Нет доступа.")
                    return
                parts = m.text.split()
                if len(parts) < 2:
                    self.bot.send_message(m.chat.id, "Использование: /deactivate_partner [id]")
                    return
                pid = int(parts[1])
                db.set_partner_active(pid, False)
                self.bot.send_message(m.chat.id, f"✅ Партнёр ID {pid} деактивирован.")
            except Exception as e: log.error(f"cmd_deactivate: {e}")

        @self.bot.message_handler(func=lambda m: True)
        def fallback(m):
            self.bot.send_message(m.chat.id, "Напишите /start чтобы начать поиск 👟\nИли /admin для администраторов.")

    def run(self):
        self.reviews.register()
        self.admin.register()
        self.vendor.register()
        self.client.register()
        self._setup_commands()
        self.scheduler.start()
        log.info("Jarvis bot started!")
        self.bot.polling(none_stop=True, timeout=60, long_polling_timeout=60)


if __name__ == "__main__":
    JarvisBot().run()
